from __future__ import annotations

import hashlib
import json
import os
import re
import time
import threading
import getpass
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Optional, Dict, List, Any

import chromadb
from chromadb.utils.embedding_functions import SentenceTransformerEmbeddingFunction
from rapidfuzz import fuzz

from utils.logger import get_logger
from utils.config import get_config
from core.session_activity import session_activity
from core.feedback import FeedbackStore
from core.brain import session_state

logger = get_logger("personal_rag")

# PRE-FIX AUDIT (read before changes):
# 1. Minimum score and related RAG thresholds now come from config.rag.
# 2. _boost_filename_matches: called in PersonalRAGIndex.search() after vector scoring.
# 3. RAG injection: pipeline._get_rag_context -> augmented_command + llm_router indexed_context.
# 4. Reranking: none before this pass (vector + fuzz boost + threshold only).
# 5. Chunk size/overlap: config rag.chunk_size=600, rag.chunk_overlap=100
#    (StructureAwareChunker uses MAX_CHUNK_CHARS=800 for AST chunks).

# Files matching these patterns are penalized — they often contain mock/fixture data.
DEPRIORITIZED_FILE_PATTERNS = [
    "test_",
    "_test.py",
    "_spec.py",
    "spec_",
    "mock",
    "fixture",
    "conftest",
    "rag_diagnostic",
    "voice_command_harness",
    "runtime_spot_check",
]

RERANKER_MODEL_DEFAULT = "cross-encoder/ms-marco-MiniLM-L-6-v2"

SUPPORTED_TEXT_EXTENSIONS = {
    ".txt", ".md", ".py", ".json", ".yaml", ".yml",
    ".csv", ".ini", ".cfg", ".toml", ".log",
    ".js", ".ts", ".html", ".css", ".sh", ".bat",
    ".rs", ".go", ".java", ".c", ".cpp", ".h",
}

CODE_EXTENSIONS = {
    ".py", ".js", ".ts", ".java", ".c", ".cpp", ".h",
    ".rs", ".go", ".sh", ".bat", ".css", ".html",
}

# Binary/junk extensions to always skip
SKIP_EXTENSIONS = {
    ".pyc", ".pyo", ".exe", ".dll", ".so", ".dylib",
    ".zip", ".tar", ".gz", ".bz2", ".7z", ".rar",
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".ico", ".webp",
    ".mp3", ".mp4", ".wav", ".avi", ".mkv", ".flac",
    ".whl", ".egg", ".db", ".sqlite", ".sqlite3",
    ".bin", ".dat", ".lock", ".map",
}


@dataclass
class RagSearchHit:
    """Normalized search result used for boosting, penalties, and reranking."""

    path: str
    content: str
    score: float
    title: str = ""
    kind: str = "document"
    metadata: dict = field(default_factory=dict)
    rerank_score: float | None = None

    @property
    def filename(self) -> str:
        return self.path or self.title or ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": self.path,
            "text": self.content,
            "content": self.content,
            "title": self.title or os.path.basename(self.path),
            "kind": self.kind,
            "score": self.score,
            "rerank_score": self.rerank_score,
            "file_extension": self.metadata.get("file_extension", ""),
            "parent_folder": self.metadata.get("parent_folder", ""),
            "metadata": dict(self.metadata),
        }

    @classmethod
    def from_payload(cls, payload: dict[str, Any]) -> "RagSearchHit":
        path = payload.get("path") or payload.get("metadata", {}).get("filepath") or ""
        return cls(
            path=path,
            content=payload.get("content") or payload.get("text") or "",
            score=float(payload.get("score", 0.0)),
            title=payload.get("title") or os.path.basename(path),
            kind=payload.get("kind", "document"),
            metadata=dict(payload.get("metadata") or {}),
            rerank_score=payload.get("rerank_score"),
        )


@dataclass
class RagChunk:
    source_path: str
    text: str
    chunk_index: int
    file_hash: str
    modified_at: float
    title: str = ""
    kind: str = "document"
    importance: int = 0
    file_extension: str = ""
    parent_folder: str = ""
    file_size_bytes: int = 0
    indexed_at: float = field(default_factory=time.time)
    chunk_type: str = "prose"
    chunk_label: str = ""
    start_line: int = 0
    language: str = "text"


class _LRUCache:
    """Simple bounded LRU cache for query results."""

    def __init__(self, maxsize: int = 200):
        self._maxsize = maxsize
        self._store: OrderedDict[str, tuple[float, list]] = OrderedDict()

    def get(self, key: str) -> tuple[float, list] | None:
        if key in self._store:
            self._store.move_to_end(key)
            return self._store[key]
        return None

    def put(self, key: str, value: tuple[float, list]) -> None:
        if key in self._store:
            self._store.move_to_end(key)
        self._store[key] = value
        while len(self._store) > self._maxsize:
            self._store.popitem(last=False)


INDEX_SCHEMA_VERSION = 3


class StructureAwareChunker:
    """Chunks documents based on structure, not arbitrary character counts."""

    MAX_CHUNK_CHARS = 800
    OVERLAP_CHARS = 150

    def chunk(self, content: str, filepath: str) -> list[dict]:
        ext = Path(filepath).suffix.lower()
        language = self._detect_language(ext)

        if language == "python":
            return self._chunk_python(content, filepath)
        if language in (
            "java", "csharp", "kotlin", "go", "cpp",
            "typescript", "javascript",
        ):
            return self._chunk_brace_language(content, filepath, language)
        if language == "markdown":
            return self._chunk_markdown(content, filepath)
        if language in ("yaml", "json", "toml"):
            return self._chunk_config(content, filepath)
        return self._chunk_prose(content, filepath)

    def _detect_language(self, ext: str) -> str:
        lang_map = {
            ".py": "python", ".java": "java", ".cs": "csharp",
            ".kt": "kotlin", ".go": "go", ".cpp": "cpp", ".c": "cpp",
            ".ts": "typescript", ".js": "javascript", ".tsx": "typescript",
            ".md": "markdown", ".yaml": "yaml", ".yml": "yaml",
            ".json": "json", ".toml": "toml", ".txt": "prose",
            ".xml": "xml", ".html": "prose", ".sql": "prose",
        }
        return lang_map.get(ext, "prose")

    def _chunk_python(self, content: str, filepath: str) -> list[dict]:
        import ast

        chunks: list[dict] = []
        try:
            tree = ast.parse(content)
            lines = content.splitlines()
            for node in ast.walk(tree):
                if isinstance(node, (ast.ClassDef, ast.FunctionDef, ast.AsyncFunctionDef)):
                    start = node.lineno - 1
                    end = getattr(node, "end_lineno", start + 50) or (start + 50)
                    chunk_text = "\n".join(lines[start:end])
                    label = (
                        f"class {node.name}"
                        if isinstance(node, ast.ClassDef)
                        else f"function {node.name}"
                    )
                    chunk_type = "class" if isinstance(node, ast.ClassDef) else "function"
                    for sub in self._split_if_large(chunk_text, self.MAX_CHUNK_CHARS):
                        chunks.append({
                            "text": sub,
                            "chunk_type": chunk_type,
                            "chunk_label": label,
                            "start_line": start + 1,
                            "filepath": filepath,
                            "language": "python",
                        })
        except SyntaxError:
            return self._chunk_prose(content, filepath)
        return chunks if chunks else self._chunk_prose(content, filepath)

    def _chunk_brace_language(self, content: str, filepath: str, language: str) -> list[dict]:
        chunks: list[dict] = []
        lines = content.splitlines()
        depth = 0
        current_block: list[str] = []
        block_start = 0
        block_label = "block"

        for i, line in enumerate(lines):
            stripped = line.strip()
            if depth == 0 and any(
                kw in stripped
                for kw in (
                    "class ", "interface ", "enum ", "struct ", "func ",
                    "public ", "private ", "protected ", "def ",
                )
            ):
                block_label = stripped[:80]
                block_start = i

            current_block.append(line)
            for ch in line:
                if ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1

            if depth == 0 and current_block:
                block_text = "\n".join(current_block)
                if len(block_text.strip()) > 20:
                    ctype = "class" if "class " in block_label else "function"
                    for sub in self._split_if_large(block_text, self.MAX_CHUNK_CHARS):
                        chunks.append({
                            "text": sub,
                            "chunk_type": ctype,
                            "chunk_label": block_label[:60],
                            "start_line": block_start + 1,
                            "filepath": filepath,
                            "language": language,
                        })
                current_block = []
                block_label = "block"

        return chunks if chunks else self._chunk_prose(content, filepath)

    def _chunk_markdown(self, content: str, filepath: str) -> list[dict]:
        sections = re.split(r"\n(?=#{1,3} )", content)
        chunks: list[dict] = []
        for section in sections:
            if not section.strip():
                continue
            label_match = re.match(r"^#{1,3} (.+)", section)
            label = label_match.group(1) if label_match else "section"
            for sub in self._split_if_large(section, self.MAX_CHUNK_CHARS):
                chunks.append({
                    "text": sub,
                    "chunk_type": "section",
                    "chunk_label": label,
                    "start_line": 0,
                    "filepath": filepath,
                    "language": "markdown",
                })
        return chunks

    def _chunk_config(self, content: str, filepath: str) -> list[dict]:
        if len(content) <= self.MAX_CHUNK_CHARS:
            return [{
                "text": content,
                "chunk_type": "config",
                "chunk_label": Path(filepath).name,
                "start_line": 1,
                "filepath": filepath,
                "language": "config",
            }]
        return self._chunk_prose(content, filepath)

    def _chunk_prose(self, content: str, filepath: str) -> list[dict]:
        paragraphs = [p.strip() for p in content.split("\n\n") if p.strip()]
        chunks: list[dict] = []
        current = ""
        for para in paragraphs:
            if len(current) + len(para) < self.MAX_CHUNK_CHARS:
                current = (current + "\n\n" + para).strip() if current else para
            else:
                if current:
                    chunks.append({
                        "text": current.strip(),
                        "chunk_type": "prose",
                        "chunk_label": Path(filepath).name,
                        "start_line": 0,
                        "filepath": filepath,
                        "language": "text",
                    })
                current = para
        if current:
            chunks.append({
                "text": current.strip(),
                "chunk_type": "prose",
                "chunk_label": Path(filepath).name,
                "start_line": 0,
                "filepath": filepath,
                "language": "text",
            })
        return chunks

    def _split_if_large(self, text: str, max_chars: int) -> list[str]:
        if len(text) <= max_chars:
            return [text]
        parts: list[str] = []
        start = 0
        while start < len(text):
            end = start + max_chars
            parts.append(text[start:end])
            if end >= len(text):
                break
            start = end - self.OVERLAP_CHARS
        return parts


def _slugify_collection_key(value: str) -> str:
    key = re.sub(r"[^a-zA-Z0-9]+", "-", (value or "").strip().lower())
    key = re.sub(r"-+", "-", key).strip("-")
    return key or "default"


@dataclass(frozen=True)
class EmbeddingProfile:
    model_name: str
    collection_key: str
    provider: str

    @classmethod
    def from_model_name(cls, model_name: str | None) -> "EmbeddingProfile":
        normalized = (model_name or "").strip()
        if not normalized or normalized.lower() in {"default", "chromadb", "chromadb-default"}:
            return cls(
                model_name="chromadb-default",
                collection_key="chromadb-default",
                provider="chromadb-default",
            )
        return cls(
            model_name=normalized,
            collection_key=_slugify_collection_key(normalized),
            provider="sentence_transformer",
        )


_EMBEDDING_FN_CACHE: Dict[tuple[str, str], SentenceTransformerEmbeddingFunction] = {}
_EMBEDDING_FN_LOCK = threading.Lock()


class HardwareEmergencyStopError(RuntimeError):
    pass


def _empty_cuda_cache() -> None:
    try:
        import torch

        torch.cuda.empty_cache()
    except Exception:
        pass


def _is_cuda_oom_exception(error: Exception) -> bool:
    try:
        import torch

        if isinstance(error, torch.cuda.OutOfMemoryError):
            return True
    except Exception:
        pass
    return "out of memory" in str(error).lower()


def _emit_hardware_emergency_stop(event_bus: Any, reason: str) -> None:
    if event_bus is None:
        return
    try:
        event_bus.emit("hardware_emergency_stop", {"reason": reason, "ts": time.time()})
    except Exception:
        logger.debug("hardware_emergency_stop_emit_failed", reason=reason, exc_info=True)


def _resolve_embedding_device(device: str | None) -> str:
    requested = (device or "").strip()
    if not requested or requested.lower() in {"auto", "default"}:
        cfg = get_config()
        hardware = getattr(cfg, "hardware", None)
        resolved = str(getattr(hardware, "embedding_device", "cpu") or "cpu").strip().lower()
        return resolved if resolved else "cpu"
    return requested


def _get_embedding_fn(model_name: str, device: str | None, event_bus: Any = None) -> SentenceTransformerEmbeddingFunction:
    resolved_device = _resolve_embedding_device(device)
    cache_key = (model_name, resolved_device)
    with _EMBEDDING_FN_LOCK:
        if cache_key not in _EMBEDDING_FN_CACHE:
            logger.info(f"Embedding model on {resolved_device}")
            logger.info("loading_embedding_model", model=model_name, device=resolved_device)
            try:
                _EMBEDDING_FN_CACHE[cache_key] = SentenceTransformerEmbeddingFunction(
                    model_name=model_name,
                    device=resolved_device,
                )
            except Exception as e:
                if resolved_device != "cpu" and _is_cuda_oom_exception(e):
                    _empty_cuda_cache()
                    logger.error("CUDA OOM — cache cleared, retrying on CPU")
                    try:
                        cache_key = (model_name, "cpu")
                        _EMBEDDING_FN_CACHE[cache_key] = SentenceTransformerEmbeddingFunction(
                            model_name=model_name,
                            device="cpu",
                        )
                    except Exception as cpu_error:
                        _emit_hardware_emergency_stop(event_bus, f"embedding_model_init:{model_name}")
                        raise HardwareEmergencyStopError(f"embedding model init failed for {model_name}") from cpu_error
                elif resolved_device != "cpu":
                    logger.warning(
                        "embedding_device_fallback",
                        model=model_name,
                        requested_device=resolved_device,
                        error=str(e),
                    )
                    fallback_device = "cpu"
                    cache_key = (model_name, fallback_device)
                    logger.info("Embedding model on cpu")
                    _EMBEDDING_FN_CACHE[cache_key] = SentenceTransformerEmbeddingFunction(
                        model_name=model_name,
                        device=fallback_device,
                    )
                else:
                    raise
            logger.info("embedding_model_loaded", model=model_name, device=cache_key[1])
        return _EMBEDDING_FN_CACHE[cache_key]




class PersonalRAGIndex:
    """Per-user RAG index with versioned collections and configurable embeddings."""

    def __init__(
        self,
        persist_directory: str,
        user_id: str = "default",
        roots: Optional[list[str]] = None,
        chunk_size: int = 600,
        chunk_overlap: int = 100,
        refresh_seconds: int = 600,
        exclude_patterns: Optional[list[str]] = None,
        collection_name: str = "dexter_personal_rag",
        embedding_model: str = "BAAI/bge-base-en-v1.5",
        embedding_device: str | None = None,
        index_schema_version: int = INDEX_SCHEMA_VERSION,
        max_context_chars: int = 3000,
        batch_size: int = 256,
        max_embedding_threads: int = 4,
        health_monitor: Any = None,
        event_bus: Any = None,
    ) -> None:
        self.user_id = (user_id or "default").lower()
        self.persist_directory = os.path.abspath(persist_directory)
        self._embedding_profile = EmbeddingProfile.from_model_name(embedding_model)
        self._index_schema_version = max(1, int(index_schema_version))
        self._collection_base_name = collection_name
        self._collection_name = (
            f"{collection_name}_idxv{self._index_schema_version}_"
            f"{self._embedding_profile.collection_key}_{self.user_id}"
        )
        self._embedding_device = embedding_device
        self._max_context_chars = max_context_chars
        self._batch_size = max(10, int(batch_size))
        self._max_embedding_threads = max(0, int(max_embedding_threads))
        self._health_monitor = health_monitor
        self._event_bus = event_bus

        # Per-user isolated chroma path
        user_path = os.path.join(self.persist_directory, f"rag_{self.user_id}")
        os.makedirs(user_path, exist_ok=True)

        self._client = chromadb.PersistentClient(path=user_path)
        collection_metadata = {
            "index_schema_version": self._index_schema_version,
            "embedding_model": self._embedding_profile.model_name,
            "embedding_provider": self._embedding_profile.provider,
            "collection_base_name": self._collection_base_name,
            "collection_name": self._collection_name,
        }
        # Create collection without an embedding function so we control batching
        # and can compute embeddings once per batch (faster, more observable).
        self._collection = self._client.get_or_create_collection(
            name=self._collection_name,
            metadata=collection_metadata,
        )
        # Preload embedding function (cached) for manual encoding
        if self._embedding_profile.provider != "chromadb-default":
            self._ef = _get_embedding_fn(self._embedding_profile.model_name, self._embedding_device, event_bus=self._event_bus)
        else:
            self._ef = None

        roots = roots or []
        self._roots = [os.path.abspath(os.path.expandvars(os.path.expanduser(r))) for r in roots if r]
        self._chunk_size = max(300, int(chunk_size))
        self._chunk_overlap = max(0, int(chunk_overlap))
        self._refresh_seconds = max(30, int(refresh_seconds))
        self._exclude_patterns = exclude_patterns or []
        self._last_refresh = 0.0
        self._last_snapshot: Dict[str, Dict[str, Any]] = {}
        self._snapshot_path = os.path.join(user_path, "snapshot.json")
        self._load_snapshot()
        self._maybe_reset_snapshot_if_empty_index()

        self._index_lock = threading.RLock()
        self._poller: Optional[threading.Thread] = None
        self._stop_poll = threading.Event()
        self._cache = _LRUCache(maxsize=200)
        self._next_poll_delay_seconds = self._refresh_seconds
        self._chunker = StructureAwareChunker()

        cfg = get_config()
        rag_cfg = cfg.rag
        self._min_score = float(rag_cfg.minimum_relevance_score)
        self._retrieval_candidates = int(rag_cfg.retrieval_candidates)
        self._final_results = int(rag_cfg.final_results)
        self._reranker_enabled = bool(rag_cfg.reranker_enabled)
        self._reranker_model = str(rag_cfg.reranker_model)
        self._background_batch_size = int(rag_cfg.background_batch_size)
        self._background_batch_sleep_seconds = float(rag_cfg.background_batch_sleep_seconds)
        self._background_embedding_threads = int(rag_cfg.background_embedding_threads)
        self._cpu_throttle_threshold_percent = int(rag_cfg.cpu_throttle_threshold_percent)
        self._vector_score_weight = float(rag_cfg.result_score_vector_weight)
        self._text_score_weight = float(rag_cfg.result_score_text_weight)
        self._importance_score_weight = float(rag_cfg.result_score_importance_weight)
        self._source_penalty_factor = float(rag_cfg.source_penalty_factor)
        self._feedback_penalty = float(rag_cfg.feedback_penalty)
        self._project_confidence_threshold = float(rag_cfg.project_confidence_threshold)
        self._filename_partial_ratio_threshold = float(rag_cfg.filename_partial_ratio_threshold)
        self._filename_fuzzy_match_threshold = float(rag_cfg.filename_fuzzy_match_threshold)
        self._filename_parent_match_threshold = float(rag_cfg.filename_parent_match_threshold)
        self._feedback_store = FeedbackStore()

        self._embed_lock = threading.Lock()
        self._indexing_active = False
        self._reranker = None
        self._reranker_ready = False
        self._reranker_init_scheduled = False
        self._startup_time = time.time()
        self._pipeline_state: str = "IDLE"
        self._last_voice_activity: float = 0.0
        self._query_cache: dict[str, tuple[list[dict], float]] = {}
        self._QUERY_CACHE_MAX_SIZE = 50
        self._QUERY_CACHE_TTL_SECONDS = 300.0

    def _emit_hardware_emergency_stop(self, reason: str) -> None:
        logger.critical("hardware_emergency_stop_request", reason=reason, user=self.user_id)
        _emit_hardware_emergency_stop(self._event_bus, reason)

    def _retry_cpu_after_cuda_oom(self, reason: str, gpu_operation, cpu_operation):
        try:
            return gpu_operation()
        except Exception as error:
            if not _is_cuda_oom_exception(error):
                raise
            _empty_cuda_cache()
            logger.error("CUDA OOM — cache cleared, retrying on CPU")
            try:
                return cpu_operation()
            except Exception as cpu_error:
                self._emit_hardware_emergency_stop(reason)
                raise HardwareEmergencyStopError(reason) from cpu_error

        logger.info(
            "personal_rag_initialized",
            user=self.user_id,
            roots=self._roots,
            collection_name=self._collection_name,
            embedding_model=self._embedding_profile.model_name,
            embedding_device=_resolve_embedding_device(self._embedding_device),
            index_schema_version=self._index_schema_version,
            chunk_size=self._chunk_size,
            batch_size=self._batch_size,
        )
        logger.info(
            "rag_thresholds_active",
            user=self.user_id,
            minimum_relevance_score=self._min_score,
            result_score_vector_weight=self._vector_score_weight,
            result_score_text_weight=self._text_score_weight,
            result_score_importance_weight=self._importance_score_weight,
            source_penalty_factor=self._source_penalty_factor,
            feedback_penalty=self._feedback_penalty,
            filename_partial_ratio_threshold=self._filename_partial_ratio_threshold,
            filename_fuzzy_match_threshold=self._filename_fuzzy_match_threshold,
            filename_parent_match_threshold=self._filename_parent_match_threshold,
            project_confidence_threshold=self._project_confidence_threshold,
            boost_cap=float(rag_cfg.boost_cap),
            refresh_idle_threshold_seconds=float(rag_cfg.refresh_idle_threshold_seconds),
        )

    # ── Polling ─────────────────────────────────────────────────────
    def start_polling(self) -> None:
        if self._poller and self._poller.is_alive():
            return
        self._stop_poll.clear()
        self._poller = threading.Thread(
            target=self._poll_loop, daemon=True, name=f"rag_poll_{self.user_id}"
        )
        self._poller.start()
        logger.info("rag_poller_started", user=self.user_id, thread=self._poller.name)

    def stop_polling(self) -> None:
        if self._poller is None:
            return
        self._stop_poll.set()
        self._poller.join(timeout=2.0)
        logger.info("rag_poller_stopped", user=self.user_id)

    def _poll_loop(self) -> None:
        while not self._stop_poll.is_set():
            try:
                self.refresh_incremental()
                self._report_health("healthy", "refresh_ok")
            except Exception as e:
                logger.warning("rag_poller_error", user=self.user_id, error=str(e), exc_info=True)
                self._report_health("degraded", str(e))
            self._stop_poll.wait(self._next_poll_delay_seconds)

    def _report_health(self, status: str, details: str = "") -> None:
        hm = self._health_monitor
        if hm is None:
            return
        try:
            if status == "healthy":
                hm.healthy("rag", details=details)
            else:
                hm.degraded("rag", details=details)
        except Exception:
            pass

    def _load_snapshot(self) -> None:
        """Load the file snapshot from disk."""
        if not os.path.exists(self._snapshot_path):
            return
        try:
            with open(self._snapshot_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                self._last_snapshot = data.get("files", {})
                self._last_refresh = data.get("last_refresh", 0.0)
            logger.info("rag_snapshot_loaded", path=self._snapshot_path, files=len(self._last_snapshot))
        except Exception as e:
            logger.warning("rag_snapshot_load_failed", error=str(e))

    def _maybe_reset_snapshot_if_empty_index(self) -> None:
        if not self._last_snapshot:
            return
        try:
            if self._collection.count() == 0:
                self._last_snapshot = {}
                self._last_refresh = 0.0
                logger.info("rag_snapshot_reset_empty_index", user=self.user_id)
        except Exception as e:
            logger.debug("rag_snapshot_count_failed", error=str(e))

    def _save_snapshot(self) -> None:
        """Save the file snapshot to disk."""
        try:
            data = {
                "files": self._last_snapshot,
                "last_refresh": self._last_refresh,
                "indexed_at": time.time(),
            }
            # Write atomically: write to temp file then replace
            tmp_path = f"{self._snapshot_path}.tmp"
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
                f.flush()
                try:
                    os.fsync(f.fileno())
                except Exception:
                    # fsync may not be available on all platforms/filesystems
                    pass
            try:
                os.replace(tmp_path, self._snapshot_path)
            except Exception:
                # Fallback to non-atomic replace
                with open(self._snapshot_path, "w", encoding="utf-8") as f2:
                    json.dump(data, f2, indent=2)

            logger.info("rag_snapshot_saved", path=self._snapshot_path, files=len(self._last_snapshot))
        except Exception as e:
            logger.warning("rag_snapshot_save_failed", error=str(e), exc_info=True)

    # ── Indexing ────────────────────────────────────────────────────
    def on_voice_activity(self) -> None:
        """Called when the user finishes an utterance — defers background refresh."""
        self._last_voice_activity = time.time()

    def set_pipeline_state(self, state: str) -> None:
        self._pipeline_state = (state or "IDLE").upper()

    def _should_refresh_now(self) -> bool:
        now = time.time()
        if not self._last_snapshot and self._last_refresh <= 0:
            return True
        if now - self._last_refresh < self._refresh_seconds:
            return False

        if self._pipeline_state not in ("IDLE", "LISTENING"):
            logger.debug(
                "rag_refresh_deferred",
                reason="pipeline_busy",
                state=self._pipeline_state,
            )
            return False

        if self._last_voice_activity and now - self._last_voice_activity < 60:
            logger.debug(
                "rag_refresh_deferred",
                reason="recent_voice_activity",
                seconds_since=round(now - self._last_voice_activity),
            )
            return False

        cfg = get_config()
        if bool(getattr(cfg.rag, "refresh_only_when_idle", True)):
            idle_threshold = float(cfg.rag.refresh_idle_threshold_seconds)
            if not session_activity.is_session_idle(idle_threshold):
                return False

        return True

    def refresh_incremental(self) -> None:
        """Walk roots, detect changed/removed files, update the index with progress tracking."""
        if not self._should_refresh_now():
            self._next_poll_delay_seconds = 60.0
            logger.info(
                "rag_refresh_deferred",
                user=self.user_id,
                reason="gating",
                retry_seconds=60,
            )
            return

        self._next_poll_delay_seconds = self._refresh_seconds
        with self._index_lock:
            t0 = time.perf_counter()
            files = []
            for root in self._roots:
                if not root or not os.path.isdir(root):
                    continue
                for path in self._walk_files(root):
                    try:
                        mtime = os.path.getmtime(path)
                        size = os.path.getsize(path)
                        files.append((path, mtime, size))
                    except Exception:
                        continue

            current_map = {p: {"mtime": m, "size": s} for p, m, s in files}
            previous_map = self._last_snapshot

            logger.info("rag_refresh_scan", user=self.user_id, found_files=len(current_map))

            to_remove = [p for p in previous_map if p not in current_map]
            to_add_or_update = [
                p for p, info in current_map.items() 
                if p not in previous_map or previous_map[p].get("mtime") != info["mtime"] or previous_map[p].get("size") != info["size"]
            ]

            if to_remove:
                for p in to_remove:
                    try:
                        self._delete_path(p)
                    except Exception as e:
                        logger.debug("rag_remove_failed", path=p, error=str(e))

            if to_add_or_update:
                chunks: List[RagChunk] = []
                file_scan_start = time.perf_counter()
                for file_num, p in enumerate(to_add_or_update, 1):
                    try:
                        text = self._read_file(p)
                        if not text.strip():
                            continue
                        file_hash = hashlib.sha256(text.encode("utf-8", errors="ignore")).hexdigest()
                        modified_at = os.path.getmtime(p)
                        title = os.path.basename(p)
                        ext = Path(p).suffix.lower()
                        parent = os.path.basename(os.path.dirname(p))
                        try:
                            fsize = os.path.getsize(p)
                        except Exception:
                            fsize = 0

                        structured = self._chunker.chunk(text, p)
                        for idx, chunk_dict in enumerate(structured):
                            chunk_text = chunk_dict.get("text", "")
                            if not chunk_text.strip():
                                continue
                            lang = chunk_dict.get("language", "text")
                            chunks.append(RagChunk(
                                source_path=p,
                                text=chunk_text,
                                chunk_index=idx,
                                file_hash=file_hash,
                                modified_at=modified_at,
                                title=title,
                                kind=self._classify_kind(ext),
                                importance=self._estimate_importance(p),
                                file_extension=ext,
                                parent_folder=parent,
                                file_size_bytes=fsize,
                                chunk_type=chunk_dict.get("chunk_type", "prose"),
                                chunk_label=chunk_dict.get("chunk_label", ""),
                                start_line=int(chunk_dict.get("start_line", 0) or 0),
                                language=lang,
                            ))

                        # Progress tracking with ETA
                        if file_num % max(1, len(to_add_or_update) // 10) == 0 or file_num == len(to_add_or_update):
                            elapsed_sec = time.perf_counter() - file_scan_start
                            progress_pct = (file_num / len(to_add_or_update)) * 100
                            rate = file_num / elapsed_sec if elapsed_sec > 0 else 0
                            remaining_files = len(to_add_or_update) - file_num
                            eta_sec = remaining_files / rate if rate > 0 else 0
                            logger.info(
                                "rag_index_progress",
                                scanned=file_num,
                                total=len(to_add_or_update),
                                pct=int(progress_pct),
                                elapsed_min=elapsed_sec / 60.0,
                                eta_min=eta_sec / 60.0,
                            )

                    except Exception as e:
                        logger.debug("rag_index_file_failed", path=p, error=str(e))

                if chunks:
                    logger.info(
                        "rag_indexing_start",
                        chunks=len(chunks),
                        batches=(len(chunks) + self._batch_size - 1) // self._batch_size,
                    )
                    self._upsert_chunks(chunks)

            self._last_snapshot = current_map
            self._last_refresh = time.time()
            self._save_snapshot()
            elapsed_ms = (time.perf_counter() - t0) * 1000
            logger.info(
                "personal_rag_incremental_refreshed",
                user=self.user_id,
                added_or_updated=len(to_add_or_update),
                removed=len(to_remove),
                total_files=len(current_map),
                elapsed_ms=f"{elapsed_ms:.0f}",
            )

    @staticmethod
    def _classify_kind(ext: str) -> str:
        if ext in CODE_EXTENSIONS:
            return "code"
        if ext in {".md", ".txt", ".log"}:
            return "text"
        if ext in {".json", ".yaml", ".yml", ".toml", ".ini", ".cfg"}:
            return "config"
        if ext in {".docx", ".pdf", ".xlsx"}:
            return "office"
        return "document"

    def _estimate_importance(self, path: str) -> int:
        name = os.path.basename(path).lower()
        score = 0
        if any(k in name for k in ("todo", "notes", "project", "readme", "meeting", "minutes", "architecture")):
            score += 20
        if "notes" in path.lower() or "projects" in path.lower():
            score += 10
        return score

    def _upsert_chunks(self, chunks: List[RagChunk]) -> None:
        if not chunks:
            return
        cfg = get_config()
        if bool(getattr(cfg.rag, "refresh_only_when_idle", True)) and self._last_snapshot:
            idle_threshold = float(cfg.rag.refresh_idle_threshold_seconds)
            if not session_activity.is_session_idle(idle_threshold):
                self._next_poll_delay_seconds = 60.0
                logger.info(
                    "rag_embedding_deferred",
                    user=self.user_id,
                    reason="session_active",
                    retry_seconds=60,
                )
                return

        self._indexing_active = True
        try:
            return self._upsert_chunks_inner(chunks)
        finally:
            self._indexing_active = False

    def _upsert_chunks_inner(self, chunks: List[RagChunk]) -> None:
        bs = min(self._batch_size, max(10, self._background_batch_size))
        total_batches = (len(chunks) + bs - 1) // bs
        logger.info("rag_upsert_start", user=self.user_id, total_chunks=len(chunks), batches=total_batches)

        # Prepare batch metadata and documents upfront
        batch_data = []
        for batch_idx in range(total_batches):
            start = batch_idx * bs
            batch = chunks[start:start + bs]
            ids = [self._chunk_id(c) for c in batch]
            documents = [c.text for c in batch]
            metadatas = [
                {
                    "path": c.source_path,
                    "filepath": c.source_path,
                    "title": c.title,
                    "kind": c.kind,
                    "chunk_index": c.chunk_index,
                    "file_hash": c.file_hash,
                    "modified_at": c.modified_at,
                    "importance": c.importance,
                    "user": self.user_id,
                    "file_extension": c.file_extension,
                    "parent_folder": c.parent_folder,
                    "file_size_bytes": c.file_size_bytes,
                    "indexed_at": c.indexed_at,
                    "chunk_type": c.chunk_type,
                    "chunk_label": c.chunk_label,
                    "start_line": c.start_line,
                    "language": c.language,
                    "index_schema_version": self._index_schema_version,
                    "embedding_model": self._embedding_profile.model_name,
                    "embedding_provider": self._embedding_profile.provider,
                    "collection_name": self._collection_name,
                }
                for c in batch
            ]
            batch_data.append((batch_idx + 1, ids, documents, metadatas))

        # Compute embeddings (serial or parallel)
        if self._ef is None:
            # Fallback: no embeddings
            for batch_idx, ids, documents, metadatas in batch_data:
                try:
                    t0 = time.perf_counter()
                    self._collection.upsert(documents=documents, metadatas=metadatas, ids=ids)
                    t1 = time.perf_counter()
                    logger.info("rag_upsert_batch_completed", user=self.user_id, batch=batch_idx, batch_size=len(documents), batch_ms=int((t1 - t0) * 1000))
                except Exception as e:
                    logger.warning("rag_upsert_batch_failed", batch=batch_idx, error=str(e), exc_info=True)
        elif self._max_embedding_threads > 1:
            # Concurrent batch embedding using ThreadPoolExecutor
            self._upsert_chunks_concurrent(batch_data)
        else:
            # Serial batch embedding (original approach)
            self._upsert_chunks_serial(batch_data)

        logger.info("rag_upsert_completed", user=self.user_id, count=len(chunks), batches=total_batches)

    def _upsert_chunks_serial(self, batch_data: List[tuple]) -> None:
        """Embed and upsert batches sequentially."""
        for batch_idx, ids, documents, metadatas in batch_data:
            if self._ef is None:
                continue
            embeddings = None
            try:
                t0 = time.perf_counter()
                embeddings = list(self._ef(documents))
                t1 = time.perf_counter()
                logger.debug("rag_batch_embedded", user=self.user_id, batch=batch_idx, batch_size=len(documents), embed_ms=int((t1 - t0) * 1000))
            except Exception as e:
                logger.warning("rag_batch_embedding_failed", batch=batch_idx, error=str(e), exc_info=True)

            try:
                t0 = time.perf_counter()
                if embeddings:
                    self._collection.upsert(documents=documents, metadatas=metadatas, ids=ids, embeddings=embeddings)
                else:
                    self._collection.upsert(documents=documents, metadatas=metadatas, ids=ids)
                t1 = time.perf_counter()
                logger.info("rag_upsert_batch_completed", user=self.user_id, batch=batch_idx, batch_size=len(documents), batch_ms=int((t1 - t0) * 1000))
            except Exception as e:
                logger.warning("rag_upsert_batch_failed", batch=batch_idx, error=str(e), exc_info=True)

    def _upsert_chunks_concurrent(self, batch_data: List[tuple]) -> None:
        """Embed batches concurrently using ThreadPoolExecutor, then upsert."""
        max_workers = max(1, min(self._max_embedding_threads, len(batch_data)))
        
        # Phase 1: Embed all batches concurrently
        embeddings_map = {}  # batch_idx -> embeddings list
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {}
            for batch_idx, ids, documents, metadatas in batch_data:
                future = executor.submit(self._embed_batch_worker, batch_idx, documents)
                futures[future] = (batch_idx, ids, documents, metadatas)

            for future in as_completed(futures):
                batch_idx, embeddings, embed_time = future.result()
                embeddings_map[batch_idx] = (embeddings, embed_time)

        # Phase 2: Upsert all batches serially (to avoid Chroma threading issues)
        for batch_idx, ids, documents, metadatas in batch_data:
            embeddings, embed_time = embeddings_map.get(batch_idx, (None, 0))
            if embeddings:
                logger.debug("rag_batch_embedded", user=self.user_id, batch=batch_idx, batch_size=len(documents), embed_ms=embed_time)

            try:
                t0 = time.perf_counter()
                if embeddings:
                    self._collection.upsert(documents=documents, metadatas=metadatas, ids=ids, embeddings=embeddings)
                else:
                    self._collection.upsert(documents=documents, metadatas=metadatas, ids=ids)
                t1 = time.perf_counter()
                logger.info("rag_upsert_batch_completed", user=self.user_id, batch=batch_idx, batch_size=len(documents), batch_ms=int((t1 - t0) * 1000))
            except Exception as e:
                logger.warning("rag_upsert_batch_failed", batch=batch_idx, error=str(e), exc_info=True)

    def _wait_if_cpu_hot(self) -> None:
        try:
            import psutil
        except ImportError:
            return

        threshold = self._cpu_throttle_threshold_percent
        while True:
            cpu = psutil.cpu_percent(interval=0.5)
            if cpu <= threshold:
                break
            logger.debug("rag_throttling", cpu_percent=cpu, threshold=threshold)
            time.sleep(3.0)

    def _embed_batch_worker(self, batch_idx: int, documents: List[str]) -> tuple:
        """Worker function for concurrent embedding with CPU throttling."""
        embeddings = None
        embed_time = 0
        try:
            self._wait_if_cpu_hot()
            try:
                import torch

                torch.set_num_threads(max(1, self._background_embedding_threads))
            except ImportError:
                pass

            def _gpu_embed():
                with self._embed_lock:
                    t0 = time.perf_counter()
                    result = list(self._ef(documents))
                    t1 = time.perf_counter()
                return result, int((t1 - t0) * 1000)

            def _cpu_embed():
                cpu_ef = _get_embedding_fn(self._embedding_profile.model_name, "cpu", event_bus=self._event_bus)
                with self._embed_lock:
                    t0 = time.perf_counter()
                    result = list(cpu_ef(documents))
                    t1 = time.perf_counter()
                return result, int((t1 - t0) * 1000)

            embeddings, embed_time = self._retry_cpu_after_cuda_oom(
                reason=f"rag_batch_embedding:{batch_idx}",
                gpu_operation=_gpu_embed,
                cpu_operation=_cpu_embed,
            )
            time.sleep(self._background_batch_sleep_seconds)
        except HardwareEmergencyStopError:
            raise
        except Exception as e:
            logger.warning("rag_batch_embedding_failed", batch=batch_idx, error=str(e), exc_info=True)
        return batch_idx, embeddings, embed_time

    def migrate_from_collection(self, source_collection_name: str, batch_size: int | None = None) -> int:
        """Re-embed stored chunks from an older collection into the current one."""
        try:
            source = self._client.get_collection(name=source_collection_name)
        except Exception as e:
            logger.info("rag_migration_source_missing", source_collection_name=source_collection_name, error=str(e))
            return 0

        payload = source.get(include=["ids", "documents", "metadatas"]) or {}
        ids = payload.get("ids") or []
        documents = payload.get("documents") or []
        metadatas = payload.get("metadatas") or []
        if not ids:
            return 0

        bs = max(1, int(batch_size or self._batch_size))
        migrated = 0
        for start in range(0, len(ids), bs):
            end = min(start + bs, len(ids))
            batch_ids = ids[start:end]
            batch_documents = documents[start:end]
            batch_metadatas = []
            for meta in metadatas[start:end]:
                meta_copy = dict(meta or {})
                meta_copy["index_schema_version"] = self._index_schema_version
                meta_copy["embedding_model"] = self._embedding_profile.model_name
                meta_copy["embedding_provider"] = self._embedding_profile.provider
                meta_copy["migrated_at"] = time.time()
                meta_copy["migrated_from_collection"] = source_collection_name
                batch_metadatas.append(meta_copy)

            try:
                self._collection.upsert(documents=batch_documents, metadatas=batch_metadatas, ids=batch_ids)
                migrated += len(batch_ids)
            except Exception as e:
                logger.warning("rag_migration_batch_failed", source_collection_name=source_collection_name, batch_start=start, error=str(e))

        logger.info(
            "rag_migration_completed",
            source_collection_name=source_collection_name,
            destination_collection_name=self._collection_name,
            migrated=migrated,
        )
        return migrated

    def migrate_from_model(self, source_model: str, source_index_schema_version: int | None = None) -> int:
        source_profile = EmbeddingProfile.from_model_name(source_model)
        schema_version = self._index_schema_version if source_index_schema_version is None else max(1, int(source_index_schema_version))
        source_collection_name = (
            f"{self._collection_base_name}_idxv{schema_version}_"
            f"{source_profile.collection_key}_{self.user_id}"
        )
        return self.migrate_from_collection(source_collection_name)

    def _delete_path(self, path: str) -> None:
        try:
            payload = self._collection.get(include=["metadatas"]) or {}
            ids = payload.get("ids") or []
            metadatas = payload.get("metadatas") or []
            to_delete = [ids[i] for i, m in enumerate(metadatas) if m and m.get("path") == path]
            if to_delete:
                self._collection.delete(ids=to_delete)
        except Exception as e:
            logger.debug("rag_delete_path_failed", path=path, error=str(e))

    # ── Search & Scoring ───────────────────────────────────────────
    def _cache_key(self, query: str) -> str:
        return hashlib.md5(query.lower().strip().encode("utf-8", errors="ignore")).hexdigest()

    def _get_cached(self, query: str) -> list[dict] | None:
        key = self._cache_key(query)
        entry = self._query_cache.get(key)
        if not entry:
            return None
        results, timestamp = entry
        if time.time() - timestamp < self._QUERY_CACHE_TTL_SECONDS:
            logger.debug("rag_cache_hit", query=query[:50])
            return results
        del self._query_cache[key]
        return None

    def _set_cached(self, query: str, results: list[dict]) -> None:
        if len(self._query_cache) >= self._QUERY_CACHE_MAX_SIZE:
            oldest = min(self._query_cache, key=lambda k: self._query_cache[k][1])
            del self._query_cache[oldest]
        self._query_cache[self._cache_key(query)] = (results, time.time())

    def _get_reranker(self, device: str | None = None):
        if not self._reranker_enabled:
            return None
        if not self._reranker_ready:
            elapsed = time.time() - getattr(self, "_startup_time", time.time())
            reranker_wait = float(get_config().rag.reranker_startup_wait_seconds)
            if elapsed < reranker_wait:
                if not self._reranker_init_scheduled:
                    logger.debug(
                        "rag_reranker_deferred_startup",
                        wait_seconds=round(reranker_wait - elapsed, 2),
                    )
                    self._reranker_init_scheduled = True
                return None
            self._reranker_ready = True
        if self._reranker is None:
            try:
                from sentence_transformers import CrossEncoder

                resolved_device = (device or getattr(get_config().hardware, "device", "auto") or "auto").strip().lower()
                if resolved_device in {"", "auto", "default"}:
                    resolved_device = str(getattr(get_config().hardware, "device", "cpu") or "cpu").strip().lower()
                self._reranker = CrossEncoder(self._reranker_model, max_length=512, device=resolved_device)
                logger.info("rag_reranker_loaded", model=self._reranker_model, device=resolved_device)
            except Exception as e:
                if _is_cuda_oom_exception(e) and (device or "").strip().lower() != "cpu":
                    _empty_cuda_cache()
                    logger.error("CUDA OOM — cache cleared, retrying on CPU")
                    try:
                        from sentence_transformers import CrossEncoder

                        self._reranker = CrossEncoder(self._reranker_model, max_length=512, device="cpu")
                        logger.info("rag_reranker_loaded", model=self._reranker_model, device="cpu")
                    except Exception as cpu_error:
                        self._emit_hardware_emergency_stop(f"reranker_init:{self._reranker_model}")
                        raise HardwareEmergencyStopError(f"reranker init failed for {self._reranker_model}") from cpu_error
                else:
                    logger.warning("rag_reranker_load_failed", error=str(e))
                    self._reranker_enabled = False
        return self._reranker

    def _hits_from_payload(self, payload: list[dict]) -> list[RagSearchHit]:
        return [RagSearchHit.from_payload(p) for p in payload]

    def _payload_from_hits(self, hits: list[RagSearchHit]) -> list[dict]:
        return [h.to_dict() for h in hits]

    def _apply_source_quality_penalties(self, hits: list[RagSearchHit]) -> list[RagSearchHit]:
        source_penalty_factor = float(
            getattr(self, "_source_penalty_factor", get_config().rag.source_penalty_factor)
        )
        for hit in hits:
            filename_lower = os.path.basename(hit.filename).lower()
            path_lower = hit.filename.lower()
            for pattern in DEPRIORITIZED_FILE_PATTERNS:
                if pattern in filename_lower or pattern in path_lower:
                    original = hit.score
                    hit.score = hit.score * source_penalty_factor
                    logger.debug(
                        "rag_source_penalized",
                        filename=hit.filename,
                        pattern_matched=pattern,
                        original_score=round(original, 2),
                        penalized_score=round(hit.score, 2),
                    )
                    break
        hits.sort(key=lambda h: h.score, reverse=True)
        return hits

    def _apply_feedback_penalties(self, hits: list[RagSearchHit], query: str) -> list[RagSearchHit]:
        feedback_store = getattr(self, "_feedback_store", None) or FeedbackStore()
        feedback_penalty = float(getattr(self, "_feedback_penalty", get_config().rag.feedback_penalty))
        try:
            penalized_paths = feedback_store.get_penalized_paths(query, user_scope=self.user_id)
        except Exception:
            penalized_paths = {}

        if not penalized_paths:
            return hits

        normalized_penalties = {os.path.abspath(path).lower(): count for path, count in penalized_paths.items()}
        for hit in hits:
            hit_path = os.path.abspath(hit.filename).lower() if hit.filename else ""
            if hit_path not in normalized_penalties:
                continue
            original = hit.score
            hit.score = hit.score * max(0.0, 1.0 - feedback_penalty)
            logger.debug(
                "rag_feedback_penalized",
                filename=hit.filename,
                feedback_count=normalized_penalties[hit_path],
                original_score=round(original, 2),
                penalized_score=round(hit.score, 2),
            )

        hits.sort(key=lambda h: h.score, reverse=True)
        return hits

    def _rerank_results(self, query: str, hits: list[RagSearchHit], top_n: int = 5) -> list[RagSearchHit]:
        reranker = self._get_reranker()
        if reranker is None or len(hits) <= 1:
            return hits[:top_n]

        try:
            pairs = [(query, h.content[:400]) for h in hits]
            scores = self._retry_cpu_after_cuda_oom(
                reason=f"rag_rerank:{query[:80]}",
                gpu_operation=lambda: reranker.predict(pairs),
                cpu_operation=lambda: self._get_reranker("cpu").predict(pairs),
            )
            for hit, score in zip(hits, scores):
                hit.rerank_score = float(score)

            before_top3 = [os.path.basename(h.filename) for h in hits[:3]]
            reranked = sorted(hits, key=lambda h: h.rerank_score or 0.0, reverse=True)
            after_top3 = [os.path.basename(h.filename) for h in reranked[:3]]
            logger.debug(
                "rag_reranked",
                query=query[:80],
                before_top3=before_top3,
                after_top3=after_top3,
            )
            return reranked[:top_n]
        except HardwareEmergencyStopError:
            raise
        except Exception as e:
            logger.warning("rag_rerank_failed", error=str(e))
            return hits[:top_n]

    @staticmethod
    def format_context_header(results: list[dict], query: str = "") -> str:
        code_extensions = {
            ".py", ".java", ".js", ".ts", ".php", ".html", ".css",
            ".jsx", ".tsx", ".go", ".rs", ".cpp", ".c", ".cs", ".rb",
        }
        code_count = 0
        for r in results:
            path = (r.get("path") or "").lower()
            if any(path.endswith(ext) for ext in code_extensions):
                code_count += 1

        if results and code_count >= len(results) // 2:
            return (
                "PERSONAL FILE CONTEXT — CODE FILES\n"
                "Read this code carefully. Explain what it does in plain English. "
                "Describe the logic, not just the syntax. Focus on what the user "
                "actually built and why it works the way it does.\n\n"
            )
        return (
            "PERSONAL FILE CONTEXT\n"
            "(Answer naturally as if you read these files yourself; cite by number when relevant.)\n\n"
        )

    def format_context_for_provider(
        self,
        results: list[dict],
        query: str,
        provider: str = "gemini",
    ) -> str:
        if not results:
            return ""

        if provider == "groq":
            trimmed = results[:2]
            max_chars = 250
        elif provider == "ollama":
            trimmed = results[:1]
            max_chars = 150
        else:
            trimmed = results[:3]
            max_chars = 500

        lines = [self.format_context_header(trimmed, query)]
        for i, r in enumerate(trimmed, 1):
            path = r.get("path") or ""
            fname = os.path.basename(path.replace("\\", "/"))
            folder = ""
            parts = path.replace("\\", "/").split("/")
            if len(parts) > 1:
                folder = parts[-2]
            meta = r.get("metadata") or {}
            label = meta.get("chunk_label", "")
            label_str = f" — {label}" if label else ""
            lines.append(f"[{i}] {fname}{label_str}" + (f"  ({folder})" if folder else ""))
            content = (r.get("content") or r.get("text") or "").strip()
            lines.append(content[:max_chars])
            lines.append("")

        lines.append(
            "When answering: be specific, cite [source numbers], "
            "explain what you found as if you read these files yourself."
        )
        return "\n".join(lines).strip()

    def search(self, query: str, limit: int = 4, use_cache: bool = True) -> List[Dict[str, Any]]:
        try:
            session_context = session_state.get_session_context()
            project = session_context.project
            if project and project.confidence > self._project_confidence_threshold and project.source_path:
                logger.info(
                    "session_context_loaded",
                    source="rag",
                    has_project=True,
                    confidence=round(project.confidence, 2),
                )
                project_hint = f"{project.name} {Path(project.source_path).stem}".strip()
                if project_hint and project_hint.lower() not in (query or "").lower():
                    query = f"{project_hint} {query}".strip()
        except Exception:
            pass

        query_key = f"{self.user_id}:{query}:{limit}"
        now = time.time()

        if use_cache:
            session_cached = self._get_cached(query)
            if session_cached is not None:
                return session_cached
            lru_cached = self._cache.get(query_key)
            if lru_cached and now - lru_cached[0] < 60.0:
                return lru_cached[1]

        try:
            n_candidates = max(limit, self._retrieval_candidates)
            query_embeddings = None
            if self._ef is not None:
                try:
                    def _gpu_query_embeddings():
                        with self._embed_lock:
                            t0 = time.perf_counter()
                            result = list(self._ef([query]))
                            t1 = time.perf_counter()
                        logger.debug(
                            "rag_query_embedded",
                            user=self.user_id,
                            embed_ms=int((t1 - t0) * 1000),
                            indexing_active=self._indexing_active,
                        )
                        return result

                    def _cpu_query_embeddings():
                        cpu_ef = _get_embedding_fn(self._embedding_profile.model_name, "cpu", event_bus=self._event_bus)
                        with self._embed_lock:
                            return list(cpu_ef([query]))

                    query_embeddings = self._retry_cpu_after_cuda_oom(
                        reason=f"rag_query_embedding:{query[:80]}",
                        gpu_operation=_gpu_query_embeddings,
                        cpu_operation=_cpu_query_embeddings,
                    )
                except HardwareEmergencyStopError:
                    raise
                except Exception:
                    query_embeddings = None

            if query_embeddings is not None:
                results = self._collection.query(
                    query_embeddings=query_embeddings,
                    n_results=max(1, n_candidates),
                )
            else:
                results = self._collection.query(
                    query_texts=[query],
                    n_results=max(1, n_candidates),
                )

            documents = (results.get("documents") or [[]])[0]
            metadatas = (results.get("metadatas") or [[]])[0]
            distances = (results.get("distances") or [[]])[0]

            payload: List[Dict[str, Any]] = []
            for idx, document in enumerate(documents):
                meta = metadatas[idx] if idx < len(metadatas) else {}
                distance = distances[idx] if idx < len(distances) else None
                vector_score = self._distance_to_score(distance)
                text_sim = float(fuzz.partial_ratio(query, document or meta.get("title", "")))
                importance = float(meta.get("importance") or 0)
                final_score = (
                    (self._vector_score_weight * vector_score)
                    + (self._text_score_weight * text_sim)
                    + (self._importance_score_weight * importance)
                )
                path = meta.get("path", "") or meta.get("filepath", "")
                payload.append({
                    "text": document,
                    "content": document,
                    "path": path,
                    "title": meta.get("title", ""),
                    "kind": meta.get("kind", "document"),
                    "score": final_score,
                    "raw_vector_score": vector_score,
                    "file_extension": meta.get("file_extension", ""),
                    "parent_folder": meta.get("parent_folder", ""),
                    "metadata": {
                        "filepath": path,
                        "chunk_label": meta.get("chunk_label", ""),
                        "chunk_type": meta.get("chunk_type", ""),
                        "language": meta.get("language", ""),
                        "start_line": meta.get("start_line", 0),
                    },
                })

            logger.debug("rag_query_raw_results", query=query, raw_count=len(payload))

            payload = self._boost_filename_matches(payload, query)
            hits = self._hits_from_payload(payload)
            hits = self._apply_source_quality_penalties(hits)
            hits = self._apply_feedback_penalties(hits, query)

            min_score = self._min_score
            filtered_hits = [h for h in hits if h.score >= min_score]
            top_n = min(limit, self._final_results) if limit else self._final_results
            final_hits = self._rerank_results(query, filtered_hits, top_n=top_n)
            payload = self._payload_from_hits(final_hits)

            logger.debug(
                "rag_accepted_results",
                query=query,
                accepted_count=len(payload),
                threshold=min_score,
                top5=[{p.get("title") or p.get("path"): p.get("score")} for p in payload[:5]],
            )

            self._cache.put(query_key, (now, payload))
            if use_cache:
                self._set_cached(query, payload)
            logger.info(
                "rag_search",
                user=self.user_id,
                query=query,
                results=len(payload),
                threshold=min_score,
                reranked=bool(self._reranker_enabled),
            )
            return payload
        except Exception as e:
            logger.warning("personal_rag_search_failed", user=self.user_id, error=str(e), exc_info=True)
            return []

    def _boost_filename_matches(self, results: List[Dict[str, Any]], query: str) -> List[Dict[str, Any]]:
        """Apply filename/parent-folder boosting safely using metadata fields.

        Returns a new results list with scores boosted according to configured
        rules. If expected metadata keys are missing, the boost for that item is
        skipped and a warning is logged.
        """
        cfg = get_config()
        boost_cap = float(cfg.rag.boost_cap)

        def split_filename_words(text: str) -> list[str]:
            """
            Split a filename/key into words from:
            - separator tokens (hyphens/underscores/etc.)
            - CamelCase capitals
            """
            raw = (text or "").strip()
            sep_parts = [p for p in re.split(r"[\W_]+", raw) if p]
            camel_parts = re.findall(r"[A-Z]+(?=[A-Z][a-z]|[0-9]|$)|[A-Z]?[a-z]+|[0-9]+", raw)

            # Prefer CamelCase tokens when the input looks like a compound identifier.
            if len(camel_parts) > 1 and len(sep_parts) <= 1:
                return [p.lower() for p in camel_parts if p]

            # Otherwise combine separator tokens with CamelCase tokens.
            words: list[str] = []
            for p in sep_parts + camel_parts:
                w = (p or "").strip().lower()
                if not w:
                    continue
                if w not in words:
                    words.append(w)
            return words

        def abbrev_of(text: str) -> str:
            words = split_filename_words(text)
            return "".join(w[0] for w in words if w)

        boosted: List[Dict[str, Any]] = []
        for r in results:
            meta_path = r.get('path')
            title = r.get('title') or (meta_path and os.path.basename(meta_path)) or ''
            if not title and not meta_path:
                logger.warning('rag_boost_skip_missing_meta', query=query, result=str(r))
                boosted.append(r)
                continue

            filename_no_ext_raw = os.path.splitext(os.path.basename(title or meta_path))[0]
            filename_no_ext = (filename_no_ext_raw or "").lower()
            q = (query or '').lower()
            parent = (r.get('parent_folder') or '').lower()
            meta_path = (r.get("path") or "").lower()
            path_components = [c for c in re.split(r"[\\/]+", meta_path) if c]
            total_bonus = 0.0

            # Filename partial match
            try:
                pr = fuzz.partial_ratio(q, filename_no_ext)
            except Exception:
                pr = 0
            if pr >= cfg.rag.filename_partial_ratio_threshold:
                total_bonus += 15

            # Abbreviation of filename (e.g., ORS -> office-reporting-system)
            try:
                if filename_no_ext:
                    abbrev = abbrev_of(filename_no_ext_raw)
                    # Query may contain arbitrary spacing; compare on a compacted string.
                    q_compact = re.sub(r"[^a-z0-9]+", "", q)
                    if abbrev and abbrev in q_compact:
                        total_bonus += 12
            except Exception:
                pass

            # Parent directory match
            try:
                if path_components:
                    best = 0
                    for comp in path_components:
                        best = max(best, fuzz.partial_ratio(q, comp))
                    if best >= cfg.rag.filename_parent_match_threshold:
                        total_bonus += 10
                else:
                    parent_score = fuzz.partial_ratio(q, parent) if parent else 0
                    if parent_score >= cfg.rag.filename_parent_match_threshold:
                        total_bonus += 10
            except Exception:
                pass

            # Directory name heuristic
            try:
                keywords = ('projects', 'documents', 'source', 'src', 'code')
                if path_components and any(any(k in comp for k in keywords) for comp in path_components):
                    total_bonus += 5
                elif parent and any(k in parent for k in keywords):
                    total_bonus += 5
            except Exception:
                pass

            # camelCase / PascalCase detection: check if filename tokens appear concatenated
            try:
                # Use the original-cased filename for CamelCase splitting.
                tokens = split_filename_words(filename_no_ext_raw)
                q_words = set(re.findall(r"[a-z0-9]+", q))
                if tokens and all(tok in q_words for tok in tokens):
                    total_bonus += 10
            except Exception:
                pass

            # Cap bonus
            if total_bonus > boost_cap:
                total_bonus = boost_cap

            if total_bonus > 0:
                r['score'] = float(r.get('score', 0.0)) + float(total_bonus)
            boosted.append(r)

        boosted.sort(key=lambda x: x.get('score', 0.0), reverse=True)
        return boosted

    def build_context(self, query: str, limit: int = 4, summary: bool = False) -> str:
        """Build LLM context from search results with structure-aware source headers."""
        results = self.search(query, limit=limit)
        if not results:
            return ""

        cfg = get_config()
        excerpt_max = int(cfg.rag.excerpt_max_chars)
        lines = [
            "PERSONAL FILE CONTEXT",
            f"(These are actual excerpts from the user's files, retrieved for query: '{query}')",
            "Answer naturally from this content. Cite source numbers when relevant.",
            "",
        ]

        seen_keys: set[str] = set()
        idx = 0
        code_langs = {
            "python", "java", "csharp", "kotlin", "go",
            "cpp", "typescript", "javascript",
        }

        for result in results:
            meta = result.get("metadata") or {}
            filepath = meta.get("filepath") or result.get("path") or "unknown"
            chunk_label = meta.get("chunk_label", "")
            chunk_type = meta.get("chunk_type", "")
            language = meta.get("language", "")

            content = (result.get("content") or result.get("text") or "").strip()
            if self._is_import_only(content) or not content:
                continue

            dedupe_key = f"{filepath}:{chunk_label}:{content[:80]}"
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)
            idx += 1

            display_path = PersonalRAGIndex._shorten_path(filepath)
            label_str = f" — {chunk_label}" if chunk_label else ""
            type_str = f" [{chunk_type}]" if chunk_type else ""
            lines.append(f"[{idx}] {display_path}{label_str}{type_str}")

            if language in code_langs:
                content = PersonalRAGIndex._remove_import_lines(content)
                snippet = content[:excerpt_max]
                lines.append(f"```{language}")
                lines.append(snippet)
                lines.append("```")
            else:
                flat = " ".join(content.split())
                if len(flat) > excerpt_max:
                    flat = flat[: excerpt_max - 3].rstrip() + "..."
                lines.append(flat)
            lines.append("")

        if idx == 0:
            return ""

        lines.append(
            "When answering: be specific, cite [source numbers], "
            "explain what you found as if you read these files yourself."
        )
        return "\n".join(lines).strip()

    def reindex_all(self) -> None:
        """Drop and rebuild the index with structure-aware chunks."""
        with self._index_lock:
            try:
                all_data = self._collection.get(include=["metadatas"]) or {}
                ids = all_data.get("ids") or []
                if ids:
                    self._collection.delete(ids=ids)
            except Exception as e:
                logger.warning("rag_reindex_clear_failed", error=str(e))

            self._last_snapshot = {}
            for root in self._roots:
                if not root or not os.path.isdir(root):
                    continue
                for path in self._walk_files(root):
                    try:
                        mtime = os.path.getmtime(path)
                        self._last_snapshot[path] = {"mtime": mtime, "size": os.path.getsize(path)}
                    except Exception:
                        continue
            self.refresh_incremental()
            logger.info("rag_reindex_all_completed", user=self.user_id)

    @staticmethod
    def _shorten_path(filepath: str) -> str:
        parts = Path(filepath).parts
        return "/".join(parts[-3:]) if len(parts) >= 3 else filepath

    @staticmethod
    def _remove_import_lines(code: str) -> str:
        lines = code.splitlines()
        prefixes = ("import ", "package ", "using ", "#include", "require(")
        meaningful = [
            ln for ln in lines
            if not any(ln.strip().startswith(kw) for kw in prefixes)
        ]
        return "\n".join(meaningful)

    def get_all_indexed_filenames(self) -> list[str]:
        """
        Return all indexed document paths from Chroma metadata only.
        No embedding / vector comparison is performed.
        """
        try:
            payload = self._collection.get(include=["metadatas"]) or {}
            metadatas = payload.get("metadatas") or []
            paths: list[str] = []
            for meta in metadatas:
                if isinstance(meta, dict):
                    p = meta.get("path") or ""
                    if p:
                        paths.append(str(p))
            return paths
        except Exception as e:
            logger.warning("personal_rag_all_filenames_failed", error=str(e), exc_info=True)
            return []

    @staticmethod
    def _is_import_only(text: str) -> bool:
        """
        True if the snippet contains only import/package/include/using directives.
        Used to avoid injecting "structural noise" into LLM prompts.
        """
        lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
        if not lines:
            return True

        import_prefix_re = re.compile(
            r"^(import\s+|package\s+|using\s+|#include\s+|require\s+)",
            re.IGNORECASE,
        )
        return all(import_prefix_re.match(ln) for ln in lines)

    # ── File Walking ───────────────────────────────────────────────
    def _walk_files(self, root: str) -> Iterable[str]:
        for dirpath, dirnames, filenames in os.walk(root):
            # Prune excluded dirs in-place to avoid descending
            dirnames[:] = [d for d in dirnames if not self._is_excluded(os.path.join(dirpath, d))]
            for filename in filenames:
                path = os.path.join(dirpath, filename)
                if self._is_supported(path) and not self._is_excluded(path):
                    yield path

    def _is_excluded(self, path: str) -> bool:
        low = path.replace("\\", "/").lower()
        for pat in self._exclude_patterns:
            if pat.lower() in low:
                return True
        return False

    def _is_supported(self, path: str) -> bool:
        ext = Path(path).suffix.lower()
        if ext in SKIP_EXTENSIONS:
            return False
        return ext in SUPPORTED_TEXT_EXTENSIONS or ext in {".docx", ".xlsx", ".pdf"}

    # ── File Reading ───────────────────────────────────────────────
    def _read_file(self, path: str) -> str:
        ext = Path(path).suffix.lower()
        if ext == ".docx":
            return self._read_docx(path)
        if ext == ".xlsx":
            return self._read_xlsx(path)
        if ext == ".pdf":
            return self._read_pdf(path)
        with open(path, "r", encoding="utf-8", errors="replace") as handle:
            return handle.read()

    def _read_docx(self, path: str) -> str:
        try:
            from docx import Document
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            logger.debug("rag_docx_failed", path=path, error=str(e))
            return ""

    def _read_xlsx(self, path: str) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            lines: list[str] = []
            for sheet in wb.worksheets:
                lines.append(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    vals = [str(c) for c in row if c is not None]
                    if vals:
                        lines.append(" | ".join(vals))
            return "\n".join(lines)
        except Exception as e:
            logger.debug("rag_xlsx_failed", path=path, error=str(e))
            return ""

    def _read_pdf(self, path: str) -> str:
        try:
            from pypdf import PdfReader
            reader = PdfReader(path)
            pages = []
            for page in reader.pages:
                try:
                    pages.append(page.extract_text() or "")
                except Exception:
                    continue
            return "\n".join(p for p in pages if p.strip())
        except Exception as e:
            logger.debug("rag_pdf_failed", path=path, error=str(e))
            return ""

    # ── Smart Chunking ─────────────────────────────────────────────
    def _chunk_text(self, text: str, ext: str = "") -> list[str]:
        normalized = text.replace("\r\n", "\n")
        if len(normalized) <= self._chunk_size:
            return [normalized]

        if ext in CODE_EXTENSIONS:
            return self._chunk_by_lines(normalized)
        return self._chunk_by_paragraphs(normalized)

    def _chunk_by_paragraphs(self, text: str) -> list[str]:
        """Split on double-newlines (paragraphs), then merge into chunks."""
        paragraphs = re.split(r"\n\s*\n", text)
        return self._merge_segments(paragraphs)

    def _chunk_by_lines(self, text: str) -> list[str]:
        """Split code files by lines, grouping into chunk_size blocks."""
        lines = text.split("\n")
        return self._merge_segments(lines, separator="\n")

    def _merge_segments(self, segments: list[str], separator: str = "\n\n") -> list[str]:
        chunks: list[str] = []
        current: list[str] = []
        current_len = 0
        for seg in segments:
            sl = len(seg) + len(separator)
            if current_len + sl <= self._chunk_size:
                current.append(seg)
                current_len += sl
            else:
                if current:
                    chunks.append(separator.join(current).strip())
                current = [seg]
                current_len = sl
        if current:
            chunks.append(separator.join(current).strip())

        # Add overlap between chunks
        if self._chunk_overlap > 0 and len(chunks) > 1:
            merged: list[str] = [chunks[0]]
            for i in range(1, len(chunks)):
                prev = chunks[i - 1]
                overlap = prev[-self._chunk_overlap:] if len(prev) > self._chunk_overlap else prev
                merged.append((overlap + separator + chunks[i]).strip())
            chunks = merged

        return [c for c in chunks if c.strip()]

    # ── Helpers ─────────────────────────────────────────────────────
    @staticmethod
    def _chunk_id(chunk: RagChunk) -> str:
        return f"{hashlib.md5(chunk.source_path.encode('utf-8', errors='ignore')).hexdigest()}_{chunk.chunk_index}_{chunk.file_hash[:12]}"

    @staticmethod
    def _distance_to_score(distance: float | None) -> float:
        """Normalize ChromaDB distance to 0–100 using 1/(1+d) scaling."""
        if distance is None:
            return 0.0
        return 100.0 * (1.0 / (1.0 + float(distance)))


class MultiUserRAGManager:
    """Manager that returns per-user PersonalRAGIndex instances."""

    def __init__(
        self,
        persist_directory: str,
        default_roots: Optional[list[str]] = None,
        cfg: Optional[dict] = None,
        health_monitor: Any = None,
    ) -> None:
        self.persist_directory = os.path.abspath(persist_directory)
        os.makedirs(self.persist_directory, exist_ok=True)
        self._indexes: Dict[str, PersonalRAGIndex] = {}
        self._default_roots = default_roots or []
        self._cfg = cfg or {}
        self._health_monitor = health_monitor

    def get_index_for_user(self, user_id: Optional[str] = None) -> PersonalRAGIndex:
        uid = (user_id or getpass.getuser() or "default").lower()
        if uid in self._indexes:
            return self._indexes[uid]
        idx = PersonalRAGIndex(
            persist_directory=self.persist_directory,
            user_id=uid,
            roots=self._cfg.get("roots") or self._default_roots,
            chunk_size=int(self._cfg.get("chunk_size", 600)),
            chunk_overlap=int(self._cfg.get("chunk_overlap", 100)),
            refresh_seconds=int(self._cfg.get("refresh_seconds", 600)),
            exclude_patterns=self._cfg.get("exclude_patterns", []),
            embedding_model=self._cfg.get("embedding_model", "BAAI/bge-base-en-v1.5"),
            embedding_device=self._cfg.get("embedding_device"),
            index_schema_version=int(self._cfg.get("index_schema_version", INDEX_SCHEMA_VERSION)),
            max_context_chars=int(self._cfg.get("max_context_chars", 3000)),
            batch_size=int(self._cfg.get("batch_size", 256)),
            max_embedding_threads=int(self._cfg.get("max_embedding_threads", 4)),
            health_monitor=self._health_monitor,
        )
        idx.start_polling()
        self._indexes[uid] = idx
        return idx
